from openpyxl import Workbook
from openpyxl.styles import PatternFill


GREEN = PatternFill(
    start_color="90EE90",
    end_color="90EE90",
    fill_type="solid"
)

RED = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

YELLOW = PatternFill(
    start_color="FFF59D",
    end_color="FFF59D",
    fill_type="solid"
)

WHITE = PatternFill(
    start_color="FFFFFF",
    end_color="FFFFFF",
    fill_type="solid"
)


def generate_excel_report(results, output_path):

    wb = Workbook()

    ws = wb.active

    ws.title = "Comparison Report"

    headers = [
        "Day1 Log",
        "Day2 Log",
        "Change Type",
        "Similarity",
        "AI Reason"
    ]

    ws.append(headers)

    for row in results:

        ws.append([
            row["day1"],
            row["day2"],
            row["type"],
            row["score"],
            row.get("reason", "")
        ])

        current_row = ws.max_row

        if row["type"] == "Added":

            fill = GREEN

        elif row["type"] == "Removed":

            fill = RED

        elif row["type"] == "Modified":

            fill = YELLOW

        else:

            fill = WHITE

        for cell in ws[current_row]:
            cell.fill = fill

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 60

    wb.save(output_path)

    return output_path